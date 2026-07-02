from __future__ import annotations

"""Excel 导入脚本。

从本地星际战士2数据表.xlsx 导入武器图标/中文名/策略词条，输出到 store/raw/excel
与 store/catalog。excel/run.py 调用本脚本并捕获 stdout 落盘导入报告供审阅。
"""

import hashlib
import json
import re
import sys
import unicodedata
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from pipeline.common import (
    APP_ASSETS_DIR,
    PIPELINE_COLLECT_EXCEL_DIR,
    PIPELINE_STORE_CATALOG_DIR,
    PIPELINE_STORE_RAW_EXCEL_DIR,
    build_weapon_asset_path,
    load_weapon_image_name_overrides,
    read_json,
    resolve_weapon_asset_name,
    sanitize_asset_name,
    weapon_slot_directory,
    write_json,
)

WORKBOOK_FILE = PIPELINE_COLLECT_EXCEL_DIR / "星际战士2数据表.xlsx"
IMPORT_DIR = PIPELINE_STORE_RAW_EXCEL_DIR
WEAPON_IMAGE_MAP_FILE = IMPORT_DIR / "武器图片映射.json"
WEAPON_NAME_LIST_FILE = IMPORT_DIR / "武器名称清单.json"
WEAPON_NAME_MAP_EXPORT_FILE = IMPORT_DIR / "武器名称映射.json"
WEAPON_IMAGE_INDEX_FILE = IMPORT_DIR / "武器图片清单.json"
WEAPON_MANIFEST_FILE = PIPELINE_STORE_CATALOG_DIR / "武器图标清单.json"
EXCEL_EXPORT_FILE = IMPORT_DIR / "Excel原始导出.json"
STRATEGY_TERMS_FILE = IMPORT_DIR / "策略词条清单.json"
WEAPON_ICON_ROOT = Path("weapons") / "icons"
SHEET_EXPORT_FILES = {
    "主页-近战武器": IMPORT_DIR / "主页-近战武器.json",
    "天赋技能效果": IMPORT_DIR / "天赋技能效果.json",
    "围攻与策略模式属性": IMPORT_DIR / "围攻与策略模式属性.json",
}
STRATEGY_GROUP_LABELS = {
    "negative": "负面词条",
    "positive": "正面词条",
}
VERSION_PATTERN = re.compile(r"当前数据为(?P<version>\d+(?:\.\d+)?)版本")
DISPIMG_PATTERN = re.compile(r'=DISPIMG\("(?P<id>ID_[A-Z0-9]+)",\s*1\)', re.IGNORECASE)
XML_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
}
SHEET_KEYWORDS = {
    "主武器": {
        "自动爆弹步枪": "auto-bolt-rifle",
        "爆弹卡宾枪": "bolt-carbine",
        "爆弹步枪": "bolt-rifle",
        "爆弹狙击步枪": "bolt-sniper-rifle",
        "多管热熔": "multi-melta",
        "双联热熔枪": "twin-linked-melta-gun",
        "全知者爆弹卡宾枪": "occulus-bolt-carbine",
        "煽动者爆弹卡宾枪": "instigator-bolt-carbine",
        "热熔步枪": "melta-rifle",
        "激光燧发枪": "las-fusil",
        "神射手卡宾枪": "marksman-bolt-carbine",
        "重型爆弹步枪": "heavy-bolt-rifle",
        "重型爆弹枪": "heavy-bolter",
        "重型等离子焚化枪": "heavy-plasma-incinerator",
        "等离子焚化枪": "plasma-incinerator",
        "焚焰枪": "pyreblaster",
        "焚焰炮": "pyrecannon",
        "追猎者爆弹步枪": "stalker-bolt-rifle",
    },
    "副武器": {
        "爆弹手枪": "bolt-pistol",
        "重爆弹手枪": "heavy-bolt-pistol",
        "地狱火手枪": "inferno-pistol",
        "高能爆燃手枪": "neo-volkite-pistol",
        "等离子手枪": "plasma-pistol",
        "单手爆弹卡宾枪": "bolt-carbine-one-handed",
    },
}
WEAPON_TITLE_COLUMN_BY_SHEET = {
    "主武器": 4,
    "副武器": 2,
    "主页-近战武器": 4,
}


@dataclass(frozen=True)
class ImageBinary:
    ext: str
    content: bytes


@dataclass(frozen=True)
class ImportFailure:
    slug: str
    reason: str

    def to_dict(self) -> dict[str, str]:
        return {"slug": self.slug, "reason": self.reason}


@dataclass(frozen=True)
class WeaponBlock:
    display_name: str
    formula: str
    title_fill_rgb: str


CANONICAL_EXCEL_WEAPON_ITEMS: dict[str, dict[str, str]] = {
    "auto-bolt-rifle": {"excel_name": "自动爆弹步枪", "source_sheet": "主武器"},
    "bolt-carbine": {"excel_name": "爆弹卡宾枪", "source_sheet": "主武器"},
    "bolt-carbine-one-handed": {"excel_name": "单手爆弹卡宾枪", "source_sheet": "副武器"},
    "marksman-bolt-carbine": {"excel_name": "神射手卡宾枪", "source_sheet": "主武器"},
    "bolt-pistol": {"excel_name": "爆弹手枪", "source_sheet": "副武器"},
    "bolt-rifle": {"excel_name": "爆弹步枪", "source_sheet": "主武器"},
    "bolt-sniper-rifle": {"excel_name": "爆弹狙击步枪", "source_sheet": "主武器"},
    "chainsword": {"excel_name": "链锯剑", "source_sheet": "主页-近战武器"},
    "combat-knife": {"excel_name": "匕首", "source_sheet": "主页-近战武器"},
    "heavy-bolt-pistol": {"excel_name": "重爆弹手枪", "source_sheet": "副武器"},
    "heavy-bolt-rifle": {"excel_name": "重型爆弹步枪", "source_sheet": "主武器"},
    "heavy-bolter": {"excel_name": "重型爆弹枪", "source_sheet": "主武器"},
    "heavy-firearms-melee": {"excel_name": "重武器近战", "source_sheet": "主页-近战武器"},
    "heavy-plasma-incinerator": {"excel_name": "重型等离子焚化枪", "source_sheet": "主武器"},
    "inferno-pistol": {"excel_name": "地狱火手枪", "source_sheet": "副武器"},
    "instigator-bolt-carbine": {"excel_name": "煽动者爆弹卡宾枪", "source_sheet": "主武器"},
    "las-fusil": {"excel_name": "激光燧发枪", "source_sheet": "主武器"},
    "melta-rifle": {"excel_name": "热熔步枪", "source_sheet": "主武器"},
    "multi-melta": {"excel_name": "多管热熔", "source_sheet": "主武器"},
    "neo-volkite-pistol": {"excel_name": "高能爆燃手枪", "source_sheet": "副武器"},
    "occulus-bolt-carbine": {"excel_name": "全知者爆弹卡宾枪", "source_sheet": "主武器"},
    "omnissiah-axe": {"excel_name": "欧姆尼赛亚战斧", "source_sheet": "主页-近战武器"},
    "plasma-incinerator": {"excel_name": "等离子焚化枪", "source_sheet": "主武器"},
    "plasma-pistol": {"excel_name": "等离子手枪", "source_sheet": "副武器"},
    "power-axe": {"excel_name": "动力斧", "source_sheet": "主页-近战武器"},
    "power-fist": {"excel_name": "动力拳", "source_sheet": "主页-近战武器"},
    "power-sword": {"excel_name": "动力剑", "source_sheet": "主页-近战武器"},
    "pyreblaster": {"excel_name": "焚焰枪", "source_sheet": "主武器"},
    "pyrecannon": {"excel_name": "焚焰炮", "source_sheet": "主武器"},
    "stalker-bolt-rifle": {"excel_name": "追猎者爆弹步枪", "source_sheet": "主武器"},
    "thunder-hammer": {"excel_name": "雷霆锤", "source_sheet": "主页-近战武器"},
    "twin-linked-melta-gun": {"excel_name": "双联热熔枪", "source_sheet": "主武器"},
}
HERO_TITLE_FILL_RGB = "FF7030A0"
EXCEL_NON_HERO_TITLE_EXCEPTIONS: set[str] = set()
EXCLUDED_CANONICAL_WEAPON_SLUGS = {"twin-linked-melta-gun"}


def _is_excluded_weapon_block_name(display_name: str) -> bool:
    normalized = str(display_name or "").strip()
    return normalized.startswith("英雄")


def _normalize_fill_rgb(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text if len(text) == 8 else ""


def _is_hero_weapon_block(*, display_name: str, title_fill_rgb: str) -> bool:
    if str(display_name or "").strip() in EXCEL_NON_HERO_TITLE_EXCEPTIONS:
        return False
    if _normalize_fill_rgb(title_fill_rgb) == HERO_TITLE_FILL_RGB:
        return True
    return _is_excluded_weapon_block_name(display_name)


def _normalize_text(value: str) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip().lower()


def _ensure_formula(value: Any) -> str:
    return str(value or "").strip().replace("=_xlfn.", "=")


def _extract_dispimg_id(formula: str) -> str:
    match = DISPIMG_PATTERN.search(_ensure_formula(formula))
    return match.group("id") if match else ""


def _clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("=_xlfn.", "=")


def _trim_row(values: list[str]) -> list[str]:
    trimmed = list(values)
    while trimmed and not str(trimmed[-1]).strip():
        trimmed.pop()
    return trimmed


def _worksheet_target(worksheet: Any) -> str:
    raw_path = str(getattr(worksheet, "path", "") or "").lstrip("/")
    return raw_path.removeprefix("xl/") if raw_path else ""


def export_workbook_raw() -> list[dict[str, Any]]:
    """从标准 Excel 重建可审阅 raw JSON。

    这里保留公式文本，尤其是 WPS 的 DISPIMG 公式，后续武器图标导入会从该导出中
    对照图片 ID。导出时只裁掉每行尾部空白，避免把格式化过的 XFD 列全部写入仓库。
    """

    workbook = load_workbook(WORKBOOK_FILE, data_only=False)
    payload: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in worksheet.iter_rows(max_row=worksheet.max_row, max_col=worksheet.max_column):
            trimmed = _trim_row([_clean_cell_value(cell.value) for cell in row])
            rows.append(trimmed)
        while rows and not rows[-1]:
            rows.pop()

        entry = {
            "sheet_name": worksheet.title,
            "target": _worksheet_target(worksheet),
            "rows": rows,
        }
        payload.append(entry)
        export_path = SHEET_EXPORT_FILES.get(worksheet.title)
        if export_path:
            write_json(export_path, entry)

    write_json(EXCEL_EXPORT_FILE, payload)
    return payload


def _strategy_title(value: Any) -> str:
    title = str(value or "").split("：", 1)[0].strip()
    return re.sub(r"^\[[^\]]+\]", "", title).strip()


def _strategy_term_key_title(value: Any) -> str:
    text = str(value or "").strip()
    title = _strategy_title(text)
    if title == "危险环境":
        if "泰伦" in text:
            return "危险环境（泰伦）"
        if "混沌" in text:
            return "危险环境（混沌）"
    if title == "战斗精通":
        if "近战增强" in text:
            return "战斗精通（近战）"
        if "远程增强" in text:
            return "战斗精通（远程）"
    return title


def _extract_workbook_versions(export_payload: list[dict[str, Any]]) -> dict[str, str]:
    versions: dict[str, str] = {}
    for entry in export_payload:
        if not isinstance(entry, dict):
            continue
        sheet_name = str(entry.get("sheet_name", "")).strip()
        for row in entry.get("rows", []):
            if not isinstance(row, list):
                continue
            for cell in row:
                match = VERSION_PATTERN.search(str(cell or ""))
                if match:
                    versions[sheet_name] = match.group("version")
                    break
            if sheet_name in versions:
                break
    return versions


def _extract_strategy_terms(rows: list[list[str]], excluded_texts: set[str] | None = None) -> dict[str, list[str]]:
    """从策略 sheet 文本行抽取正/负词条。

    excluded_texts 为应屏蔽的词条文本（Excel 里灰色字体标记的未实装/废弃词条），
    命中则跳过，不进入词条库。颜色信息在 export_workbook_raw 已丢失，故由
    export_strategy_terms 用 openpyxl 单独读取后传入。
    """
    excluded = excluded_texts or set()
    groups = {key: [] for key in STRATEGY_GROUP_LABELS}
    seen = {key: set() for key in STRATEGY_GROUP_LABELS}
    active_group = ""

    for row in rows:
        cells = [str(cell or "").strip() for cell in row if str(cell or "").strip()]
        if STRATEGY_GROUP_LABELS["negative"] in cells:
            active_group = "negative"
        if STRATEGY_GROUP_LABELS["positive"] in cells:
            active_group = "positive"
        if active_group not in groups:
            continue
        for cell in cells:
            if cell in STRATEGY_GROUP_LABELS.values() or "：" not in cell:
                continue
            if cell in excluded:
                continue
            if cell in seen[active_group]:
                continue
            seen[active_group].add(cell)
            groups[active_group].append(cell)
    return groups


GREYED_STRATEGY_FONT_RGBS = {
    "FF666666",
    "FF767171",
    "FF7F7F7F",
    "FF808080",
    "FF999999",
    "FFA6A6A6",
}


def _is_greyed_strategy_font(color: Any) -> bool:
    """仅识别维护约定中的显式 RGB 灰色，避免 theme/auto/indexed 误杀。"""
    if color is None or getattr(color, "type", None) != "rgb":
        return False
    rgb = str(getattr(color, "rgb", "") or "").upper()
    return rgb in GREYED_STRATEGY_FONT_RGBS


def _collect_greyed_strategy_texts() -> set[str]:
    """读取策略 sheet 中灰色字体单元格的文本，作为应屏蔽词条。

    Excel 维护者用灰色字体标记未实装/废弃词条（如 FF767171），这些不应进入
    词条库。export_workbook_raw 只保留纯文本丢失颜色，故在此用 openpyxl 单独读。
    只识别显式 RGB 灰色；theme/auto/indexed 与黑色都按正常词条处理。
    """
    greyed: set[str] = set()
    workbook = load_workbook(WORKBOOK_FILE, data_only=False)
    if "围攻与策略模式属性" not in workbook.sheetnames:
        return greyed
    worksheet = workbook["围攻与策略模式属性"]
    for row in worksheet.iter_rows():
        for cell in row:
            text = str(cell.value or "").strip()
            if not text or "：" not in text:
                continue
            color = cell.font.color if cell.font and cell.font.color else None
            if _is_greyed_strategy_font(color):
                greyed.add(text)
    return greyed


def export_strategy_terms(export_payload: list[dict[str, Any]]) -> dict[str, Any]:
    strategy_sheet = next(
        (
            entry
            for entry in export_payload
            if isinstance(entry, dict) and str(entry.get("sheet_name", "")).strip() == "围攻与策略模式属性"
        ),
        {"rows": []},
    )
    greyed_texts = _collect_greyed_strategy_texts()
    groups = _extract_strategy_terms(strategy_sheet.get("rows", []), excluded_texts=greyed_texts)
    versions = _extract_workbook_versions(export_payload)
    items = [*groups["negative"], *groups["positive"]]
    payload = {
        "source_sheet": "围攻与策略模式属性",
        "source_version": versions.get("围攻与策略模式属性", ""),
        "versions_by_sheet": versions,
        "items": items,
        "groups": groups,
        "group_counts": {key: len(value) for key, value in groups.items()},
        "greyed_excluded": sorted(greyed_texts),
        "titles": {
            key: [_strategy_term_key_title(term) for term in value]
            for key, value in groups.items()
        },
    }
    write_json(STRATEGY_TERMS_FILE, payload)
    return payload


def _load_workbook_sheet_rows() -> dict[str, list[list[str]]]:
    payload = read_json(EXCEL_EXPORT_FILE, [])
    sheet_rows: dict[str, list[list[str]]] = {}
    for entry in payload:
        if not isinstance(entry, dict):
            continue
        sheet_name = str(entry.get("sheet_name", "")).strip()
        rows = entry.get("rows", [])
        if sheet_name and isinstance(rows, list):
            sheet_rows[sheet_name] = rows
    return sheet_rows


def _find_formula_in_export(sheet_rows: dict[str, list[list[str]]], source_sheet: str, excel_name: str) -> str:
    rows = sheet_rows.get(source_sheet, [])
    for row_index, row in enumerate(rows):
        if not isinstance(row, list):
            continue
        for column_index, cell in enumerate(row):
            if str(cell).strip() != excel_name:
                continue
            for neighbor_index in (column_index - 1, column_index + 1):
                if 0 <= neighbor_index < len(row):
                    formula = _ensure_formula(row[neighbor_index])
                    if _extract_dispimg_id(formula):
                        return formula
            for scan_row in (row_index - 1, row_index + 1):
                if 0 <= scan_row < len(rows) and isinstance(rows[scan_row], list):
                    neighbor_row = rows[scan_row]
                    for neighbor_index in (column_index, column_index - 1, column_index + 1):
                        if 0 <= neighbor_index < len(neighbor_row):
                            formula = _ensure_formula(neighbor_row[neighbor_index])
                            if _extract_dispimg_id(formula):
                                return formula
    return ""


def _load_openpyxl_rows(source_sheet: str) -> list[dict[int, str]]:
    workbook = load_workbook(WORKBOOK_FILE, data_only=False)
    target_sheet = None
    for sheet_name in workbook.sheetnames:
        if _normalize_text(sheet_name) == _normalize_text(source_sheet):
            target_sheet = workbook[sheet_name]
            break
    if target_sheet is None:
        return []
    rows: list[dict[int, str]] = []
    for row in target_sheet.iter_rows():
        row_map = {int(cell.column): "" if cell.value is None else str(cell.value) for cell in row}
        if row_map:
            rows.append(row_map)
    return rows


def _collect_weapon_blocks(source_sheet: str) -> list[WeaponBlock]:
    blocks: list[WeaponBlock] = []
    workbook = load_workbook(WORKBOOK_FILE, data_only=False)
    if source_sheet not in workbook.sheetnames:
        return blocks
    worksheet = workbook[source_sheet]
    title_column = WEAPON_TITLE_COLUMN_BY_SHEET.get(source_sheet, 4)
    for row_index in range(1, worksheet.max_row + 1):
        formula = _ensure_formula(worksheet.cell(row=row_index, column=1).value)
        title_cell = worksheet.cell(row=row_index, column=title_column)
        display_name = str(title_cell.value or "").strip()
        title_fill_rgb = _normalize_fill_rgb(title_cell.fill.fgColor.rgb)
        if _extract_dispimg_id(formula) and display_name and not _is_hero_weapon_block(display_name=display_name, title_fill_rgb=title_fill_rgb):
            blocks.append(WeaponBlock(display_name=display_name, formula=formula, title_fill_rgb=title_fill_rgb))
    return blocks


def _build_slug_formula_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for source_sheet, keyword_map in SHEET_KEYWORDS.items():
        for block in _collect_weapon_blocks(source_sheet):
            normalized_name = _normalize_text(block.display_name)
            for keyword, slug in keyword_map.items():
                if _normalize_text(keyword) == normalized_name:
                    lookup[slug] = {
                        "formula": block.formula,
                        "source_sheet": source_sheet,
                        "excel_name": block.display_name,
                    }
                    break
    return lookup


def _discover_new_weapon_blocks() -> list[dict[str, Any]]:
    """发现 Excel 中 canonical 白名单之外的武器块，标记待审新增。

    Excel 维护者新增武器行时，这些行不在 CANONICAL_EXCEL_WEAPON_ITEMS 内，过去
    会被静默忽略。这里按 source_sheet 汇总 canonical 已知武器名，遍历全部武器
    sheet（主/副/近战），收集白名单外的块，以稳定 slug(excel-discovered-{hash})
    纳入候选，带 pending_review 标记，靠 diff 审阅。中文名无法 slugify，故用
    sha1 前 8 位保证跨 run 稳定。正式纳入运行数据仍需人工补 canonical + 职业池映射。
    """
    known_names_by_sheet: dict[str, set[str]] = {}
    for metadata in CANONICAL_EXCEL_WEAPON_ITEMS.values():
        sheet = str(metadata.get("source_sheet", "")).strip()
        name = _normalize_text(metadata.get("excel_name", ""))
        if sheet and name:
            known_names_by_sheet.setdefault(sheet, set()).add(name)
    discovered: list[dict[str, Any]] = []
    seen_slugs: set[str] = set()
    for source_sheet in WEAPON_TITLE_COLUMN_BY_SHEET:
        for block in _collect_weapon_blocks(source_sheet):
            normalized_name = _normalize_text(block.display_name)
            if not normalized_name or normalized_name in known_names_by_sheet.get(source_sheet, set()):
                continue
            # 过滤纯数字/纯符号噪声块（序号图片等），武器名必含中文或字母
            if not re.search(r"[一-鿿 a-zA-Z]", block.display_name):
                continue
            digest = hashlib.sha1(f"{source_sheet}:{block.display_name}".encode("utf-8")).hexdigest()[:8]
            slug = f"excel-discovered-{digest}"
            if slug in seen_slugs:
                continue
            seen_slugs.add(slug)
            discovered.append(
                {
                    "slug": slug,
                    "excel_name": block.display_name,
                    "source_sheet": source_sheet,
                    "image_formula": block.formula,
                    "review_reason": "excel_new_row",
                    "pending_review": True,
                }
            )
    return discovered


def _validate_workbook_structure(export_payload: list[dict[str, Any]]) -> list[dict[str, str]]:
    """校验工作簿表名/表头，产出 warning 列表（不阻断导入）。

    检查两类可靠信号：期望的武器/数据 sheet 是否齐全；武器 sheet 是否仍能抓到
    至少一个武器块（标题列错位或表结构变化时抓到 0 块）。反向枚举未识别 sheet
    噪声大（WPS 工作簿天然含 WpsReserved_* 与辅助 sheet），故不采用。warning
    只落盘 excel_import_report.json，不阻断导入。
    """
    warnings: list[dict[str, str]] = []
    payload_sheets = {
        str(entry.get("sheet_name", "")).strip()
        for entry in export_payload
        if isinstance(entry, dict)
    }
    expected_sheets = set(SHEET_EXPORT_FILES) | set(WEAPON_TITLE_COLUMN_BY_SHEET)
    for sheet in sorted(expected_sheets):
        if sheet not in payload_sheets:
            warnings.append(
                {
                    "code": "missing_expected_sheet",
                    "sheet": sheet,
                    "message": f"期望的 sheet 缺失：{sheet}，Excel 表名可能被改动",
                }
            )
    for source_sheet in WEAPON_TITLE_COLUMN_BY_SHEET:
        if source_sheet not in payload_sheets:
            continue
        if not _collect_weapon_blocks(source_sheet):
            warnings.append(
                {
                    "code": "title_column_misaligned",
                    "sheet": source_sheet,
                    "message": f"{source_sheet} 未能抓到任何武器块，标题列可能错位或表结构变化",
                }
            )
    return warnings


def _build_image_rel_lookup() -> dict[str, str]:
    with zipfile.ZipFile(WORKBOOK_FILE) as workbook_zip:
        rels_root = ElementTree.fromstring(workbook_zip.read("xl/_rels/cellimages.xml.rels"))
    lookup: dict[str, str] = {}
    for relation in rels_root:
        relation_id = relation.attrib.get("Id", "")
        target = relation.attrib.get("Target", "")
        if relation_id and target:
            normalized_target = target.removeprefix("/")
            lookup[relation_id] = normalized_target if normalized_target.startswith("xl/") else f"xl/{normalized_target}"
    return lookup


def _build_dispimg_binary_lookup() -> dict[str, ImageBinary]:
    rel_lookup = _build_image_rel_lookup()
    with zipfile.ZipFile(WORKBOOK_FILE) as workbook_zip:
        cellimages_root = ElementTree.fromstring(workbook_zip.read("xl/cellimages.xml"))
        lookup: dict[str, ImageBinary] = {}
        for cell_image in cellimages_root.findall("etc:cellImage", XML_NS):
            name_node = cell_image.find(".//xdr:cNvPr", XML_NS)
            dispimg_id = "" if name_node is None else name_node.attrib.get("name", "")
            blip = cell_image.find(".//a:blip", XML_NS)
            embed_id = "" if blip is None else blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "")
            target = rel_lookup.get(embed_id, "")
            if not dispimg_id or not target:
                continue
            try:
                content = workbook_zip.read(target)
            except KeyError:
                continue
            lookup[dispimg_id] = ImageBinary(ext=Path(target).suffix.lower() or ".png", content=content)
    return lookup


def _append_missing_manifest_items(items: list[dict[str, Any]], manifest: dict[str, Any]) -> list[dict[str, Any]]:
    existing_slugs = {str(item.get("slug", "")).strip() for item in items}
    appended = list(items)
    for entry in manifest.get("weapons", []):
        if not isinstance(entry, dict):
            continue
        slug = str(entry.get("slug", "")).strip()
        if slug.startswith("excel-discovered-"):
            continue
        if not slug or slug in existing_slugs or slug in EXCLUDED_CANONICAL_WEAPON_SLUGS:
            continue
        appended.append({
            "slug": slug,
            "excel_name": "",
            "image_formula": f"__MISSING_MAPPING__:{slug}",
            "source_sheet": "",
        })
    return appended


def _is_stable_weapon_item(item: dict[str, Any]) -> bool:
    slug = str(item.get("slug", "")).strip()
    return bool(slug) and not item.get("pending_review") and not slug.startswith("excel-discovered-")


def _build_manifest_lookup(manifest: dict[str, Any]) -> dict[str, str]:
    return {
        str(entry.get("slug", "")).strip(): str(entry.get("asset_path", "")).strip()
        for entry in manifest.get("weapons", [])
        if isinstance(entry, dict)
    }


def _infer_sheet_for_slug(slug: str) -> str:
    metadata = CANONICAL_EXCEL_WEAPON_ITEMS.get(str(slug or "").strip(), {})
    return str(metadata.get("source_sheet", "")).strip()


def _slot_type_from_source_sheet(source_sheet: str) -> str:
    normalized = str(source_sheet or "").strip()
    if normalized == "主武器":
        return "primary"
    if normalized == "副武器":
        return "secondary"
    if normalized == "主页-近战武器":
        return "melee"
    return ""


def _find_formula_for_item(item: dict[str, Any], sheet_rows: dict[str, list[list[str]]], slug_formula_lookup: dict[str, dict[str, str]]) -> str:
    current_formula = _ensure_formula(item.get("image_formula", ""))
    lookup_entry = slug_formula_lookup.get(str(item.get("slug", "")).strip(), {})
    lookup_formula = _ensure_formula(lookup_entry.get("formula", ""))
    if _extract_dispimg_id(lookup_formula):
        return lookup_formula
    source_sheet = str(item.get("source_sheet", "")).strip()
    excel_name = str(item.get("excel_name", "")).strip()
    if source_sheet and excel_name:
        formula = _find_formula_in_export(sheet_rows, source_sheet, excel_name)
        if formula:
            return formula
    return current_formula


def _resolve_item_asset_fields(
    *,
    slug: str,
    source_sheet: str,
    excel_name: str,
    default_name: str,
    overrides: dict[str, dict[str, Any]],
) -> tuple[str, str, str]:
    slot_directory = weapon_slot_directory(source_sheet=source_sheet)
    asset_name = resolve_weapon_asset_name(
        slug=slug,
        excel_name=excel_name,
        default_name=default_name,
        overrides=overrides,
    )
    asset_path = build_weapon_asset_path(source_sheet=source_sheet, asset_name=asset_name)
    return slot_directory, asset_name, asset_path


def _materialize_items(
    image_map: dict[str, Any],
    manifest: dict[str, Any],
    sheet_rows: dict[str, list[list[str]]],
    slug_formula_lookup: dict[str, dict[str, str]],
    overrides: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[ImportFailure]]:
    # 过滤掉历史落盘的待审新增项，避免它们在 WEAPON_IMAGE_MAP_FILE / manifest
    # 中累积后被 _append_missing_manifest_items 反复加回。按 slug 前缀排除最稳妥
    # （历史项可能丢失 pending_review 字段）；每次由 _discover_new_weapon_blocks
    # 重新决定哪些新行进入待审。
    raw_base_items = list(image_map.get("items", [])) if isinstance(image_map, dict) else []
    base_items = [
        item
        for item in raw_base_items
        if isinstance(item, dict)
        and not item.get("pending_review")
        and not str(item.get("slug", "")).startswith("excel-discovered-")
    ]
    items = _append_missing_manifest_items(base_items, manifest)
    for slug, metadata in CANONICAL_EXCEL_WEAPON_ITEMS.items():
        if slug in EXCLUDED_CANONICAL_WEAPON_SLUGS:
            continue
        if slug not in {str(item.get("slug", "")).strip() for item in items}:
            items.append(
                {
                    "slug": slug,
                    "excel_name": metadata["excel_name"],
                    "image_formula": "",
                    "source_sheet": metadata["source_sheet"],
                }
            )
    # 追加 Excel 白名单外的新增武器行，带 pending_review 标记，靠 diff 审阅。
    existing_slugs = {str(item.get("slug", "")).strip() for item in items}
    for discovered in _discover_new_weapon_blocks():
        if discovered["slug"] in existing_slugs:
            continue
        existing_slugs.add(discovered["slug"])
        items.append(discovered)
    finalized: list[dict[str, Any]] = []
    failures: list[ImportFailure] = []
    for item in items:
        slug = str(item.get("slug", "")).strip()
        if slug in EXCLUDED_CANONICAL_WEAPON_SLUGS:
            continue
        canonical_metadata = CANONICAL_EXCEL_WEAPON_ITEMS.get(slug, {})
        source_sheet = str(canonical_metadata.get("source_sheet", "")).strip() or str(item.get("source_sheet", "")).strip() or _infer_sheet_for_slug(slug)
        excel_name = str(canonical_metadata.get("excel_name", "")).strip() or str(item.get("excel_name", "")).strip()
        formula = _find_formula_for_item({**item, "source_sheet": source_sheet, "excel_name": excel_name}, sheet_rows, slug_formula_lookup)
        default_name = str(item.get("asset_file_name", "")).removesuffix(".png").strip()
        slot_directory, asset_name, asset_path = _resolve_item_asset_fields(
            slug=slug,
            source_sheet=source_sheet,
            excel_name=excel_name,
            default_name=default_name,
            overrides=overrides,
        )
        finalized_item = {
            **item,
            "excel_name": excel_name,
            "source_sheet": source_sheet,
            "image_formula": formula,
            "asset_file_name": f"{sanitize_asset_name(asset_name)}.png",
            "asset_path": asset_path,
            "directory_label": slot_directory,
        }
        finalized.append(finalized_item)
        if not _extract_dispimg_id(formula):
            failures.append(ImportFailure(slug=slug, reason="缺少可解析的 DISPIMG 公式"))
    return finalized, failures


def _write_png(target_path: Path, image_binary: ImageBinary) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(BytesIO(image_binary.content)) as image:
        image.convert("RGBA").save(target_path, format="PNG")


def _cleanup_obsolete_weapon_icons(items: list[dict[str, Any]]) -> None:
    expected_paths = {
        (APP_ASSETS_DIR / Path(str(item.get("asset_path", "")).strip())).resolve()
        for item in items
        if str(item.get("asset_path", "")).strip()
    }
    icon_root = APP_ASSETS_DIR / WEAPON_ICON_ROOT
    if not icon_root.exists():
        return
    for file_path in icon_root.rglob("*.png"):
        if file_path.resolve() not in expected_paths:
            file_path.unlink()


def _import_items(items: list[dict[str, Any]], image_lookup: dict[str, ImageBinary]) -> tuple[list[str], list[ImportFailure]]:
    imported: list[str] = []
    failures: list[ImportFailure] = []
    for item in items:
        slug = str(item.get("slug", "")).strip()
        formula = _ensure_formula(item.get("image_formula", ""))
        dispimg_id = _extract_dispimg_id(formula)
        if not slug:
            continue
        # 待审新增项不导出图标、不计入 imported_count，避免污染 31 武器基线与
        # app/assets；其 has_image 信号由 excel_import_report 反映，靠 diff 审阅。
        if item.get("pending_review"):
            continue
        if not dispimg_id:
            continue
        image_binary = image_lookup.get(dispimg_id)
        if image_binary is None:
            failures.append(ImportFailure(slug=slug, reason=f"工作簿中未找到图片 ID: {dispimg_id}"))
            continue
        relative_asset_path = str(item.get("asset_path", "")).strip() or (WEAPON_ICON_ROOT / f"{slug}.png").as_posix()
        _write_png(APP_ASSETS_DIR / Path(relative_asset_path), image_binary)
        imported.append(slug)
    return imported, failures


def _dedupe_failures(failures: list[ImportFailure]) -> list[ImportFailure]:
    unique: dict[tuple[str, str], ImportFailure] = {}
    for failure in failures:
        unique[(failure.slug, failure.reason)] = failure
    return list(unique.values())


def _build_clean_excel_exports(items: list[dict[str, Any]]) -> None:
    normalized_items = [
        {
            "slug": str(item.get("slug", "")).strip(),
            "excel_name": str(item.get("excel_name", "")).strip(),
            "source_sheet": str(item.get("source_sheet", "")).strip(),
            "slot_type": _slot_type_from_source_sheet(str(item.get("source_sheet", "")).strip()),
            "directory_label": str(item.get("directory_label", "")).strip(),
            "asset_file_name": str(item.get("asset_file_name", "")).strip(),
            "asset_path": str(item.get("asset_path", "")).strip(),
        }
        for item in items
        if _is_stable_weapon_item(item) and str(item.get("excel_name", "")).strip()
    ]

    normalized_items.sort(key=lambda item: (item["slot_type"], item["excel_name"], item["slug"]))

    write_json(
        WEAPON_NAME_MAP_EXPORT_FILE,
        {
            "items": [
                {
                    "slug": item["slug"],
                    "excel_name": item["excel_name"],
                    "slot_type": item["slot_type"],
                    "source_sheet": item["source_sheet"],
                }
                for item in normalized_items
            ],
            "notes": "Excel 规范武器命名。英雄级/变体不会额外生成独立武器分类；运行层统一使用基础武器名。",
        },
    )
    write_json(
        WEAPON_NAME_LIST_FILE,
        {
            "items": [
                {
                    "slug": item["slug"],
                    "excel_name": item["excel_name"],
                    "slot_type": item["slot_type"],
                }
                for item in normalized_items
            ],
            "notes": "去重后的 Excel 武器清单，仅保留运行层使用的基础武器名。",
        },
    )
    write_json(
        WEAPON_IMAGE_INDEX_FILE,
        {
            "items": [
                {
                    "slug": item["slug"],
                    "excel_name": item["excel_name"],
                    "slot_type": item["slot_type"],
                    "source_sheet": item["source_sheet"],
                    "directory_label": item["directory_label"],
                    "asset_file_name": item["asset_file_name"],
                    "asset_path": item["asset_path"],
                }
                for item in normalized_items
            ],
            "notes": "Excel 图像索引。英雄级/词条说明不单列为武器项，图片路径按运行期规范输出。",
        },
    )
    write_json(
        WEAPON_MANIFEST_FILE,
        {
            "weapons": [
                {
                    "slug": item["slug"],
                    "image_key": f"weapon_{item['slug']}_img",
                    "asset_path": item["asset_path"],
                }
                for item in normalized_items
                if item["asset_path"]
            ]
        },
    )


def import_weapon_icons() -> dict[str, Any]:
    export_payload = export_workbook_raw()
    strategy_payload = export_strategy_terms(export_payload)
    image_map = read_json(WEAPON_IMAGE_MAP_FILE, {"items": []})
    manifest = read_json(WEAPON_MANIFEST_FILE, {"weapons": []})
    sheet_rows = _load_workbook_sheet_rows()
    slug_formula_lookup = _build_slug_formula_lookup()
    image_lookup = _build_dispimg_binary_lookup()
    overrides = load_weapon_image_name_overrides()

    finalized_items, mapping_failures = _materialize_items(image_map, manifest, sheet_rows, slug_formula_lookup, overrides)
    stable_items = [item for item in finalized_items if _is_stable_weapon_item(item)]
    imported_slugs, import_failures = _import_items(stable_items, image_lookup)
    _cleanup_obsolete_weapon_icons(stable_items)
    write_json(WEAPON_IMAGE_MAP_FILE, {"items": stable_items})
    _build_clean_excel_exports(stable_items)

    failures = _dedupe_failures([*mapping_failures, *import_failures])
    discovered_new_items = [
        {
            "slug": str(item.get("slug", "")).strip(),
            "excel_name": str(item.get("excel_name", "")).strip(),
            "source_sheet": str(item.get("source_sheet", "")).strip(),
            "review_reason": str(item.get("review_reason", "")).strip(),
            "has_image": bool(_extract_dispimg_id(_ensure_formula(item.get("image_formula", "")))),
        }
        for item in finalized_items
        if item.get("pending_review")
    ]
    return {
        "workbook": WORKBOOK_FILE.relative_to(PROJECT_ROOT).as_posix(),
        "imported_count": len(imported_slugs),
        "imported_slugs": imported_slugs,
        "failure_count": len(failures),
        "failures": [failure.to_dict() for failure in failures],
        "strategy_term_count": len(strategy_payload.get("items", [])),
        "strategy_group_counts": strategy_payload.get("group_counts", {}),
        "excel_source_version": strategy_payload.get("source_version", ""),
        "versions_by_sheet": strategy_payload.get("versions_by_sheet", {}),
        "greyed_excluded": strategy_payload.get("greyed_excluded", []),
        "discovered_new_items": discovered_new_items,
        "header_warnings": _validate_workbook_structure(export_payload),
        "output_dir": (APP_ASSETS_DIR / WEAPON_ICON_ROOT).relative_to(PROJECT_ROOT).as_posix(),
    }


if __name__ == "__main__":
    print(json.dumps(import_weapon_icons(), ensure_ascii=False, indent=2))
